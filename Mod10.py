#!/usr/bin python3
from ast import Num
import email
from fileinput import filename
from multiprocessing import dummy
import numbers
from sqlite3 import Date
import sys
from tokenize import Number
from turtle import clear
from colorama import Fore, init, Back, Style
import openpyxl
from openpyxl.styles import NamedStyle
import re
from codicefiscale import codicefiscale
import requests
from requests.exceptions import HTTPError
#import setuptools
#from setuptools import setup
from datetime import datetime


#versione = "0.2.0 del "+str(datetime.today().strftime('%d-%m-%Y'))
versione = "2." + str(datetime.today().strftime('%m%d%S'))+" - (c) "+str(datetime.today().strftime('%Y'))+" - A.I.A.C. Service s.r.l. Unipersonale "

#__annotations__="Controllo i dati prima di inviarli al Settore Tecnico FIGC"
#__package__= "Corretto errore di eliminazione apice"


def release () :
    # Visualizzo la release alla partenza
    #print(Back.BLUE,Fore.WHITE,"Mod10 Rel."+__version__+" - "+__annotations__,Back.RESET+'\n')
    print(Back.BLUE,Fore.WHITE,"Mod10 Rel."+versione,Back.RESET+'\n')

def hello (nome) :
    # Comunico quale file sto leggendo il numero di righe e colonne
    #print(Back.BLUE,Fore.WHITE,'Leggo il file:',nome,Back.RESET+'\n')
    print(Fore.GREEN,'# Leggo il file:',nome,Back.RESET+' #')
    
def printErrore (errdescr) :
    print(Back.RED+Fore.YELLOW+errdescr+Back.RESET+Fore.RESET)

def checknome (nome) :
    # Ritorno la stringa pulita da caratteri che non ci devono essere
    nome = nome.replace("'"," ")
    nome = nome.replace("à", "a")
    nome = nome.replace("è", "e")
    nome = nome.replace("é", "e")
    nome = nome.replace("ì", "i")
    nome = nome.replace("ò", "o")
    nome = nome.replace("ù", "u")
    nome = nome.replace("À", "A")
    nome = nome.replace("È", "E")
    nome = nome.replace("Ì", "I")
    nome = nome.replace("Ò", "O")
    nome = nome.replace("Ù", "U")
    return nome.upper()

def charexactly (nome, numchar) :
    # Controllo che la stringa passata abbia l'esatto numero di caratteri ad esempio CAP di 5 chr
    conta = len(nome)
    if (conta > numchar or conta < numchar) :
        return False
    return True
    
def stringoverflow (nome, numchar) :
    conta = len(nome)
    #print(nome,numchar,conta)
    if (conta > numchar) :
        return True
    return False

regex = re.compile(r'([A-Za-z0-9]+[.-_])*[A-Za-z0-9]+@[A-Za-z0-9-]+(\.[A-Z|a-z]{2,})+')

def isValidMail(emailaddress):
    # controllo che sia un indirizzo email scritto correttamente
    if re.fullmatch(regex, emailaddress):
      return True
    else:
      return False

def checksesso (name) :
    # Controllo cosa è inserito nel campo sesso
    if(name == 'M' or name == 'F') :
        return True
    else : 
        return False
    
def checkStudio (name) :
    # Controlllo che il titoli di studio sia tra quelli previsti dal modulo
    if name == 'LM':
        return True
    elif name == 'IP':
        return True
    elif name == 'DM':
        return True
    elif name == 'IS':
        return True
    elif name == 'ME':
        return True
    elif name == 'AC':
        return True
    elif name == 'EA':
        return True
    else:
        return False
        
def checkindirizzo (nome) :
    # Controllo che il campo indirizzo contenga un indicazione corretta
    nome = checknome(nome)
    if nome.find("VIA") >= 0 : return True
    if nome.find("VIALE") >= 0 : return True
    if nome.find("LARGO") >= 0 : return True
    if nome.find("PIAZZA") >= 0 : return True
    if nome.find("CORSO") >= 0 : return True
    if nome.find("CONTRADA") >= 0 : return True
    if nome.find("VICOLO") >= 0 : return True
    if nome.find("LOCALITA") >= 0 : return True
    if nome.find("STRADA") >= 0 : return True
    
    return False
    
def checktelefono (nome) :
    # elimino i campi in esubero dal numero di cellulare
    nome = nome.replace("+39","")
    nome = nome.replace("+", "")
    nome = nome.replace(".", "")
    nome = nome.replace(" ", "")
    nome = nome.replace("-", "")
    nome = nome.replace("_", "")
    return nome

def checkvoto (voto):
    # Controllo che il valore del voto porti alla promozione altrimenti lo segnalo
    if 84 <= voto <= 140:
        return True
    else: 
        return False


def urlcomunenascita(cod_catasto) :
    # controllo il comune di nascita risalendo dal codice del catasto del codice fiscale sopratutto per inserire la nazione in caso di stranieri
    try:
        parametri = {"token": "05052021", "cod": cod_catasto}
        r = requests.post("http://segreteria.assoallenatori.it/api/catasto.php", data=parametri)
        return r.text.lstrip().upper()
        #print(r.json())
    except HTTPError as http_err:
        print(f'HTTP error occurred: {http_err}')
    except Exception as err:
        print(f'Other error occurred: {err}')

def urlprovincianascita(cod_catasto) :
    # controllo il comune di nascita risalendo dal codice del catasto del codice fiscale sopratutto per inserire la nazione in caso di stranieri
    try:
        parametri = {"token": "05052021", "cod": cod_catasto}
        r = requests.post("http://segreteria.assoallenatori.it/api/cf_prov.php", data=parametri)
        return r.text.lstrip().upper()
        #print(r.json())
    except HTTPError as http_err:
        print(f'HTTP error occurred: {http_err}')
    except Exception as err:
        print(f'Other error occurred: {err}')

def urlprovinciaresidenza(comune) :
    # controllo il comune di residenza risalendo dal nome del comune così riesco a controllare l'esatta definizione del comune e la provincia corretta
    try:
        parametri = {"token": "05052021", "cod": comune}
        r = requests.post("http://segreteria.assoallenatori.it/api/local_prov.php", data=parametri)
        return r.text.lstrip().upper()
        #print(r.json())
    except HTTPError as http_err:
        print(f'HTTP error occurred: {http_err}')
    except Exception as err:
        print(f'Other error occurred: {err}')

errore = False
errdescr = ''
#init(convert=True)
print("\n")
release()
#path = input("Inserisci il nome del file xls, ad ex- 70304.xlsx : ")
if len(sys.argv) < 2 :
    namefile = input("Inserisci il nome del Modello 10 in formato Excel : ")
else:
    namefile = sys.argv[1];     

# Controllo che il nome abbia l'estensione xlsx altrimenti la aggiungo
num_ext = namefile.find('.xlsx')
#print('result = '+str(num_ext))
if (num_ext < 0) : namefile = namefile + '.xlsx'

# Creo il file di salvataggio dei dati
savefile = namefile.replace(".xlsx", "_checked.xlsx")
#input_col_name = input("Enter colname, ex- Endpoint : ")
try:
    #release()
        
    print(Fore.RESET)
    #Apro il file Excel
    wb_obj = openpyxl.load_workbook(namefile)
    
    #leggi i nomi dei fogli presenti nel file excel
    sheetname = wb_obj.sheetnames
    #for i in range(0, len(sheetname)):
    #    print(Fore.BLUE + sheetname[i])

    #sheet = wb_obj.get_sheet_by_name(sheetname[0])
    sheet = wb_obj[sheetname[0]]
    # Conto Righe e Colonne
    max_column=sheet.max_column
    max_row=sheet.max_row
    
    hello(namefile+' - Colonne: '+str(max_column)+" / "+'Righe: '+str(max_row))
    #print(Fore.GREEN + 'Colonne: '+str(max_column))
    #print(Fore.RED + 'Righe: '+str(max_row))
    
    # Leggo riga per riga partendo dalla seconda lascio stare le intestazioni
    for i in range(2, max_row+1):
        #errore = False
        riga = str(i)
        
        # COGNOME
        cognome = sheet['A'+str(i)].value
        sheet['A'+str(i)] = checknome(cognome)        
        # NOME
        nome = sheet['B'+str(i)].value
        sheet['B'+str(i)] = checknome(nome)
        # LUOGO DI NASCITA
        luogonascita = sheet['C'+str(i)].value
        #sheet['C'+str(i)] = checknome(luogonascita)
        # PROVINCIA DI NASCITA
        prvnascita = sheet['D'+str(i)].value
        if not charexactly(prvnascita,2) : 
            errore = True
            errdescr = errdescr + riga + "->Errore Provincia Nascita: " + prvnascita + "\n "
        # DATA DI NASCITA
        datanascita = sheet['E'+str(i)].value
        celldate = sheet['E'+str(i)]
        #print(celldate.style)
        nsddmmyyyy=NamedStyle(name="ddmmaaaa-"+str(i), number_format="DD/MM/YYYY")
        if(celldate.style == "Normale"):
            celldate.style = nsddmmyyyy
        
        #INDIRIZZO        
        indirizzo = sheet['F'+str(i)].value
        #controllo che l'indirizzo sia completo di VIA PIAZZA ecc. ecc. 
        if not checkindirizzo(indirizzo) : 
            # Non essendo un errore vincolante tolgo il blocco
            #errore = True
            #errdescr = errdescr + riga + "->Errore Controllare Indirizzo: " + indirizzo + "\n "
            printErrore(riga + "->Errore Controllare Indirizzo: " + indirizzo)
        #controllo che non superi i 30 caratteri
        if stringoverflow(indirizzo, 30) : 
            errore = True
            errdescr = errdescr + riga + "->Errore Lunghezza indirizzo: " + indirizzo + "\n"
        #se l'indirizzo è corretto lo copio
        sheet['F'+str(i)] = checknome(indirizzo)
        # CAP    
        cap = str(sheet['G'+str(i)].value)
        if not charexactly(str(cap),5) : 
            #errore = True
            #errdescr = errdescr + riga + "->Errore CAP: " + cap + "\n "
            printErrore(riga + "->Errore Controllare CAP: " + cap)
        # LUOGO DI RESIDENZA    
        luogoresidenza = sheet['H'+str(i)].value
        sheet['H'+str(i)] = checknome(luogoresidenza)
        
        # PROVINCIA DI RESIDENZA
        provincia = sheet['I'+str(i)].value
        if not charexactly(provincia,2) : 
            errore = True
            errdescr = errdescr + riga + "->Errore Provincia Residenza: " + provincia + "\n "
        else :
            prvres = urlprovinciaresidenza(luogoresidenza)
            if (prvres == "ZZ") :
                errore = True
                errdescr = errdescr + riga + "->Provincia di Residenza NON trovata " + luogoresidenza + "\n"
            else :            
                sheet['I'+str(i)] = prvres
        
        # TELEFONO FISSO
        # Azzero il telefono fisso che non serve più
        teldummy = sheet['J'+str(i)].value
        sheet['J'+str(i)] = ""
        
        # TELEFONO CELLULARE
        cellulare = str(sheet['K'+str(i)].value)
        sheet['K'+str(i)] = checktelefono(cellulare)
        
        #INDIRIZZO EMAIL
        emailaddress = sheet['L'+str(i)].value
        if not isValidMail(emailaddress) :
            errore = True
            errdescr = errdescr + riga + "->Errore Indirizzo EMail: " + emailaddress + "\n "
        
        # SESSO    
        sesso = sheet['O'+str(i)].value
        if not checksesso(sesso) :
            errore = True
            errdescr = errdescr + riga + "->Errore Sesso: " + sesso + "\n "
        
        # CODICE FISCALE
        CodiceFIscaleFile = sheet['M'+str(i)].value
        CFValid = codicefiscale.is_valid(CodiceFIscaleFile)
        if(not CFValid) : 
            errore = True
            #errdescr = errdescr + riga + "->Controllare Codice Fiscale: " + str(CodiceFIscaleFile) + " il calcolo a me risulta: " + str(CFCalcolate) + "\n "
            errdescr = errdescr + riga + "->Controllare Codice Fiscale: " + str(CodiceFIscaleFile) + "\n "
        else:
            # prendo il codice catastale dal codice fiscale e lo passo alla procedura
            catasto = CodiceFIscaleFile[11:15]
            comune = urlcomunenascita(catasto)
            if (comune == "ZZ") :
                errore = True
                errdescr = errdescr + riga + "->Comune di Nascita NON trovato " + catasto + "\n"
            else :
                sheet['C'+str(i)] = comune
            
            prov = urlprovincianascita(catasto)
            if (prov == "ZZ") :
                errore = True
                errdescr = errdescr + riga + "->Provincia di Nascita NON trovata " + catasto + "\n"
            else :            
                sheet['D'+str(i)] = prov
                #print(prov)
                        
        # TITOLO DI STUDIO   
        titolostudio = sheet['N'+str(i)].value
        if not checkStudio(titolostudio) :
            errore = True
            errdescr = errdescr + riga + "->Errore Titolo di Studio: " + titolostudio + "\n "
        # VOTO CONSEGUITO
        voto = sheet['P'+str(i)].value
        if not checkvoto(voto) :
            #errore = True
            printErrore(riga + "-> "+cognome+" "+nome+" Controllare Voto: " + str(voto))
            #errdescr = errdescr + riga + "->Controllare Voto: " + str(voto) + "\n "
    
    # FINE CICLO SULLE RIGHE
    # Visualizzo gli errori in caso ci siano
    if (errore) :
        print(Back.RED+Fore.YELLOW+errdescr+Back.RESET+Fore.RESET)
    else :
        #print(Fore.GREEN + "# " + Fore.WHITE + "Sembra che non ci siano errori rilevanti" + Fore.GREEN + " #" + Fore.RESET)
        print("\n" + Fore.WHITE + "Sembra che non ci siano errori bloccanti" + Fore.RESET + "\n")
        wb_obj.save(savefile)
        print(Fore.GREEN + "# Scritto file "+savefile+" #"+Fore.RESET + "\n\n")
    
except Exception as e:
    exception_type, exception_object, exception_traceback = sys.exc_info()
    filename = exception_traceback.tb_frame.f_code.co_filename
    line_number = exception_traceback.tb_lineno
    print(namefile,"Excel Line: ",str(i))
    print("Exception type: ", exception_type)
    print("File Name: ", filename)
    print("Code Line: ", line_number)